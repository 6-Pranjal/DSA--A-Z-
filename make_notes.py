from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── helpers ────────────────────────────────────────────────────────────────
def header_fill(color): return PatternFill("solid", fgColor=color)
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg="1F4E79", fg="FFFFFF", size=11, bold=True):
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = header_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_cell(cell, bg=None, bold=False, wrap=True, align="left"):
    cell.font = Font(bold=bold, size=10)
    if bg: cell.fill = header_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – PROGRESS TRACKER
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Progress Tracker"

headers = ["#", "Folder", "Topic", "Category", "Language", "Status", "Revision Needed?", "Date Committed"]
col_widths = [5, 32, 28, 18, 10, 12, 18, 18]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws1.cell(1, i, h)
    style_header(c)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 30

rows = [
    (1,  "1-InputOutput",                          "Basic I/O",                        "Basics",     "Java", "Done", "No",  "26-Mar-2026"),
    (2,  "2-IF_ElseIF",                            "Conditionals",                     "Basics",     "Java", "Done", "No",  "26-Mar-2026"),
    (3,  "3-Switch",                               "Switch Statement",                  "Basics",     "Java", "Done", "No",  "26-Mar-2026"),
    (4,  "4-patterns-1",                           "Pattern 1",                        "Patterns",   "Java", "Done", "No",  "29-Mar-2026"),
    (5,  "5-patterns-2",                           "Pattern 2",                        "Patterns",   "Java", "Done", "No",  "29-Mar-2026"),
    (6,  "6-patterns-3",                           "Pattern 3",                        "Patterns",   "Java", "Done", "No",  "29-Mar-2026"),
    (7,  "7-patterns-4",                           "Pattern 4",                        "Patterns",   "Java", "Done", "No",  "29-Mar-2026"),
    (8,  "8-patterns-5",                           "Pattern 5",                        "Patterns",   "Java", "Done", "No",  "29-Mar-2026"),
    (9,  "9-patterns-6",                           "Pattern 6",                        "Patterns",   "Java", "Done", "No",  "07-Apr-2026"),
    (10, "10-patterns-7",                          "Pattern 7",                        "Patterns",   "Java", "Done", "No",  "07-Apr-2026"),
    (11, "11-Math-Reverse-1",                      "Reverse Integer (LC #7)",          "Math",       "Java", "Done", "Yes", "07-Apr-2026"),
    (12, "12-Math-Pallindrom-2",                   "Palindrome Number (LC #9)",        "Math",       "Java", "Done", "No",  "07-Apr-2026"),
    (13, "13-Math-Armstrong-3",                    "Armstrong Number",                 "Math",       "Java", "Done", "No",  "07-Apr-2026"),
    (14, "14-Recurrsion-Decreasing_Increasing_num","Recursion – Print Decreasing/Inc", "Recursion",  "Java", "Done", "No",  "08-Apr-2026"),
    (15, "15-Recurrsion-fibonacci-5",              "Recursion – Fibonacci",            "Recursion",  "Java", "Done", "No",  "08-Apr-2026"),
    (16, "16-Recurrsion-sum-factorial-3-4",        "Recursion – Sum & Factorial",      "Recursion",  "Java", "Done", "Yes", "08-Apr-2026"),
    (17, "17-Recurssion-Array_Reverse-6",          "Recursion – Array Reverse",        "Recursion",  "Java", "Done", "Yes", "11-Apr-2026"),
    (18, "18-Hashing-1",                           "Hashing – Frequency Count",        "Hashing",    "Java", "Done", "No",  "17-Apr-2026"),
    (19, "19-dp-fibo-1",                           "DP – Fibonacci (Memo + Tab)",      "DP",         "Java", "Done", "Yes", "17-Apr-2026"),
    (20, "20-dp-climbingstare-2",                  "DP – Climbing Stairs",             "DP",         "Java", "Done", "Yes", "17-Apr-2026"),
    (21, "21-trees-creationand+preorderTraversal", "Tree – Creation + Preorder",       "Trees",      "Java", "Done", "No",  "19-Apr-2026"),
    (22, "22-tree-dfs-preorder-inorder-postorder", "Tree – DFS (Pre/In/Post)",         "Trees",      "Java", "Done", "No",  "19-Apr-2026"),
    (23, "23-tree-sametree",                       "Tree – Same Tree (LC #100)",       "Trees",      "Java", "Done", "Yes", "19-Apr-2026"),
    (24, "24-tree-bfs",                            "Tree – BFS Level Order",           "Trees",      "Java", "Done", "Yes", "19-Apr-2026"),
    (25, "25-tree-height",                         "Tree – Height",                    "Trees",      "Java", "Done", "No",  "19-Apr-2026"),
    (26, "26-tree-totalNodes and sum",             "Tree – Total Nodes & Sum",         "Trees",      "Java", "Done", "No",  "19-Apr-2026"),
]

cat_colors = {
    "Basics":    "D9EAD3",
    "Patterns":  "FCE5CD",
    "Math":      "FFF2CC",
    "Recursion": "CFE2F3",
    "Hashing":   "EAD1DC",
    "DP":        "D9D2E9",
    "Trees":     "C9DAF8",
}

for r, row in enumerate(rows, 2):
    bg = cat_colors.get(row[3], "FFFFFF")
    for c, val in enumerate(row, 1):
        cell = ws1.cell(r, c, val)
        style_cell(cell, bg=bg, align="center" if c in (1,5,6,7,8) else "left")
    ws1.row_dimensions[r].height = 20

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – TOPIC NOTES & KEY POINTS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Topic Notes")
headers2 = ["#", "Topic", "Category", "Key Concept / What You Did", "Important Points to Remember"]
widths2   = [5, 30, 14, 55, 65]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    c = ws2.cell(1, i, h)
    style_header(c, bg="1A5276")
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 30

notes = [
    (1,  "Basic I/O",                       "Basics",    "Used Scanner to read int and print with System.out.println()",
         "Always import java.util.*; for Scanner. Use sc.nextInt() for int, sc.next() for String."),
    (2,  "Conditionals (if/else if)",        "Basics",    "Standard if-else if-else ladder",
         "Conditions are evaluated top-down; only first true block runs."),
    (3,  "Switch Statement",                 "Basics",    "Switch-case with break",
         "Always add break; after each case or fall-through happens. Use default as catch-all."),
    (4,  "Patterns 1–7",                     "Patterns",  "Nested loops to print star/number/character patterns",
         "Outer loop = rows, inner loop = columns. Think about what changes per row before coding."),
    (5,  "Reverse Integer (LC #7)",          "Math",      "Extract digits with %10, build reversed number with *10+digit",
         "Overflow check MUST come BEFORE multiplying: check r > MAX_VALUE / 10, not after. Current code checks AFTER — BUG!"),
    (6,  "Palindrome Number (LC #9)",        "Math",      "Reverse the number and compare with original. Negative = false.",
         "Negative numbers are never palindromes. Store original before reversing."),
    (7,  "Armstrong Number",                 "Math",      "Sum of cubes of digits == original number",
         "Code hardcodes cube (d*d*d). For general n-digit Armstrong use Math.pow(d, digits)."),
    (8,  "Recursion – Decreasing/Increasing","Recursion", "Print n down to 1 (decreasing). Move print AFTER recursive call for increasing.",
         "Print BEFORE call = decreasing. Print AFTER call = increasing. Base case: n==0 return."),
    (9,  "Recursion – Fibonacci",            "Recursion", "fib(n) = fib(n-1) + fib(n-2), base: fib(0)=0, fib(1)=1",
         "Pure recursion is O(2^n). Always upgrade to memoization for large n."),
    (10, "Recursion – Sum & Factorial",      "Recursion", "sum(n) = n + sum(n-1), base: n==1 return 1",
         "Factorial base case should be n==0 return 1 (not n==1) to handle 0! = 1. Commented factorial line left in code."),
    (11, "Recursion – Array Reverse",        "Recursion", "Print arr[n] then recurse with n-1; prints in reverse",
         "This prints reverse but does NOT modify the array in-place. For in-place reverse use two-pointer swap approach."),
    (12, "Hashing – Frequency Count",        "Hashing",   "HashMap<Integer,Integer> to count occurrences of each element",
         "Use getOrDefault(key, 0)+1 as a cleaner alternative to containsKey check. Iteration via entrySet()."),
    (13, "DP – Fibonacci (Memo + Tab)",      "DP",        "Memoization: top-down recursion + dp array. Tabulation: bottom-up loop.",
         "Tabulation loop goes i=2 to n (inclusive). fibo_tab returns dp[n-1] — off-by-one risk, verify indexing carefully."),
    (14, "DP – Climbing Stairs",             "DP",        "Ways to climb n stairs (1 or 2 steps) = fib(n+1). Memo + pure recursion.",
         "BUG: climbingStairs_mem(n) called without dp[] argument in main. Also 'return=' is a syntax error in climbingStairs_rec."),
    (15, "Tree – Creation + Preorder",       "Trees",     "Build binary tree from int[] using recursion. -1 = null node.",
         "Static index i must be reset to -1 before each tree creation. Preorder: Root → Left → Right."),
    (16, "Tree – DFS (Pre/In/Post)",         "Trees",     "Three DFS traversals implemented recursively.",
         "Preorder: Root-L-R | Inorder: L-Root-R | Postorder: L-R-Root. Inorder of BST gives sorted output."),
    (17, "Tree – Same Tree (LC #100)",       "Trees",     "Serialize both trees into lists (preorder with nulls), compare lists.",
         "BUG: l.add(null) — ArrayList<Integer> can hold null but equals() comparison still works. Cleaner approach: direct recursive comparison without lists."),
    (18, "Tree – BFS Level Order",           "Trees",     "Queue-based BFS. Use null as level separator in queue.",
         "BUG: Infinite loop if tree is non-empty — after removing null, you add null again unconditionally. Must check if queue is empty before adding null sentinel."),
    (19, "Tree – Height",                    "Trees",     "height = max(height(left), height(right)) + 1. Base: null → 0.",
         "Height = number of nodes on longest path. If counting edges instead, return -1 for null. Know which definition is asked."),
    (20, "Tree – Total Nodes & Sum",         "Trees",     "nodeCount = 1 + count(left) + count(right). sum = data + sum(left) + sum(right).",
         "Both follow same post-order pattern. Base case: root==null return 0."),
]

for r, row in enumerate(notes, 2):
    bg = cat_colors.get(row[2], "FFFFFF")
    for c, val in enumerate(row, 1):
        cell = ws2.cell(r, c, val)
        style_cell(cell, bg=bg, align="center" if c == 1 else "left")
    ws2.row_dimensions[r].height = 55

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – BUGS & ERRORS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Bugs & Errors")
headers3 = ["#", "File / Topic", "Bug / Error", "Severity", "Fix / Correct Approach"]
widths3   = [5, 30, 55, 12, 60]
for i, (h, w) in enumerate(zip(headers3, widths3), 1):
    c = ws3.cell(1, i, h)
    style_header(c, bg="7B241C")
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 30

sev_colors = {"Critical": "F4CCCC", "High": "FCE5CD", "Medium": "FFF2CC", "Low": "D9EAD3"}

bugs = [
    (1,  "20-dp-climbingstare-2",        "Syntax error: 'return=' is not valid Java. Should be 'return climbingStairs_rec(n-1)+climbingStairs_rec(n-2);'",
         "Critical", "Remove '=' from return statement."),
    (2,  "20-dp-climbingstare-2",        "climbingStairs_mem(n) called in main WITHOUT passing dp[] array — compile error.",
         "Critical", "Call as climbingStairs_mem(n, dp) passing the dp array."),
    (3,  "24-tree-bfs",                  "Infinite loop: null sentinel is added back unconditionally after being removed, even when queue is empty.",
         "Critical", "Add: if(!q.isEmpty()) q.add(null); before adding the sentinel back. Also add a break when queue becomes empty after removing null."),
    (4,  "11-Math-Reverse-1",            "Overflow check is done AFTER r is already multiplied — check is too late and can itself overflow.",
         "High",     "Check BEFORE: if(r > Integer.MAX_VALUE/10 || r < Integer.MIN_VALUE/10) return 0;"),
    (5,  "23-tree-sametree",             "l.add(null) adds null into ArrayList<Integer>. Works but is fragile and non-standard.",
         "Medium",   "Better: directly compare recursively — if(p==null && q==null) return true; if(p==null||q==null) return false; return p.val==q.val && isSameTree(p.left,q.left) && isSameTree(p.right,q.right);"),
    (6,  "13-Math-Armstrong-3",          "Function is named 'reverse' but checks Armstrong — misleading name.",
         "Low",      "Rename to isArmstrong() for clarity."),
    (7,  "13-Math-Armstrong-3",          "Hardcoded cube (d*d*d) only works for 3-digit numbers.",
         "Medium",   "Use Math.pow(d, String.valueOf(n).length()) for general case."),
    (8,  "16-Recurrsion-sum-factorial-3-4","Factorial base case is n==1, but 0! = 1 is not handled.",
         "Medium",   "Change base case to n==0 return 1."),
    (9,  "17-Recurssion-Array_Reverse-6","Prints reverse but does not actually reverse the array in memory.",
         "Low",      "For in-place reverse use two-pointer: swap arr[start] and arr[end], recurse with start+1, end-1."),
    (10, "19-dp-fibo-1",                 "fibo_tab loop runs i<n (exclusive), so dp[n] is never filled. Returns dp[n-1] which may be wrong for edge cases.",
         "Medium",   "Loop should run i<=n and return dp[n]."),
    (11, "Folder naming",                "Typo: 'Recurrsion' (double r) and 'Recurssion' used inconsistently across folders.",
         "Low",      "Standardize to 'Recursion' in future folder names."),
    (12, "Folder naming",                "Typo: 'climbingstare' should be 'climbingStairs'.",
         "Low",      "Rename folder for clarity."),
]

for r, row in enumerate(bugs, 2):
    bg = sev_colors.get(row[3], "FFFFFF")
    for c, val in enumerate(row, 1):
        cell = ws3.cell(r, c, val)
        style_cell(cell, bg=bg, align="center" if c in (1, 4) else "left")
    ws3.row_dimensions[r].height = 50

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – QUICK REVISION CHEATSHEET
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Revision Cheatsheet")
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 90

c = ws4.cell(1, 1, "Topic"); style_header(c, bg="145A32")
c = ws4.cell(1, 2, "Quick Formula / Template"); style_header(c, bg="145A32")
ws4.row_dimensions[1].height = 28

cheat = [
    ("Tree Creation from array",
     "i=-1 (static). i++. if arr[i]==-1 return null. Node n=new Node(arr[i]); n.left=create(arr); n.right=create(arr); return n;"),
    ("Preorder (Root-L-R)",      "print(root) → preorder(left) → preorder(right)"),
    ("Inorder (L-Root-R)",       "inorder(left) → print(root) → inorder(right)  [BST → sorted]"),
    ("Postorder (L-R-Root)",     "postorder(left) → postorder(right) → print(root)"),
    ("BFS Level Order",          "Queue q; q.add(root); q.add(null); while(!q.isEmpty()){ n=q.remove(); if(n==null){ println(); if(!q.isEmpty()) q.add(null); } else { print(n); add children; } }"),
    ("Tree Height",              "if(root==null) return 0; return Math.max(height(left), height(right)) + 1;"),
    ("Node Count",               "if(root==null) return 0; return 1 + count(left) + count(right);"),
    ("Tree Sum",                 "if(root==null) return 0; return root.data + sum(left) + sum(right);"),
    ("DP Memoization template",  "if(dp[n]!=-1) return dp[n]; ... dp[n]=solve(n-1)+solve(n-2); return dp[n];"),
    ("DP Tabulation template",   "dp[0]=base0; dp[1]=base1; for(i=2;i<=n;i++) dp[i]=dp[i-1]+dp[i-2]; return dp[n];"),
    ("HashMap frequency count",  "map.put(key, map.getOrDefault(key,0)+1);"),
    ("Reverse integer safely",   "while(x!=0){ if(r>MAX/10||r<MIN/10) return 0; r=r*10+x%10; x/=10; }"),
    ("Palindrome check",         "store p=x; reverse x into r; return p==r; (negative → false)"),
    ("Armstrong (3-digit)",      "sum of (each digit)^3 == original number"),
    ("Recursion – Decreasing",   "print(n); recurse(n-1);  [print BEFORE call]"),
    ("Recursion – Increasing",   "recurse(n-1); print(n);  [print AFTER call]"),
]

for r, (topic, formula) in enumerate(cheat, 2):
    c1 = ws4.cell(r, 1, topic);   style_cell(c1, bg="D5F5E3", bold=True)
    c2 = ws4.cell(r, 2, formula); style_cell(c2, bg="EAFAF1")
    ws4.row_dimensions[r].height = 40

path = r"c:\Users\pranj\OneDrive\Desktop\DSA-(A-Z)\DSA_Notes.xlsx"
wb.save(path)
print("Saved:", path)


