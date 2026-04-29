from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "DSA Notes"

# ── helpers ──────────────────────────────────────────────────────────────────
def side(): return Side(style="thin")
def border(): return Border(left=side(), right=side(), top=side(), bottom=side())

def hdr(cell, bg="1F4E79", fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()

def cell_style(cell, bg="FFFFFF", bold=False, align="left"):
    cell.font = Font(size=9, bold=bold)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    cell.border = border()

# ── columns ──────────────────────────────────────────────────────────────────
columns = [
    ("#",                    4),
    ("Folder",              28),
    ("Topic",               26),
    ("Category",            13),
    ("LeetCode #",          12),
    ("Language",            10),
    ("Status",              10),
    ("Date Committed",      15),
    ("What You Did",        40),
    ("Key Points / Formula",48),
    ("Possible Mistakes",   48),
    ("Revision Needed?",    14),
]

for col, (name, width) in enumerate(columns, 1):
    c = ws.cell(1, col, name)
    hdr(c)
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

# ── category colors ───────────────────────────────────────────────────────────
CAT = {
    "Basics":    ("D9EAD3", "274E13"),
    "Patterns":  ("FCE5CD", "7F3C00"),
    "Math":      ("FFF2CC", "7D6608"),
    "Recursion": ("CFE2F3", "1A3A5C"),
    "Hashing":   ("EAD1DC", "4A0033"),
    "DP":        ("D9D2E9", "20124D"),
    "Trees":     ("C9DAF8", "1C3A6E"),
}

# ── data ─────────────────────────────────────────────────────────────────────
# cols: #, folder, topic, category, leetcode, lang, status, date, what_you_did, key_points, mistakes, revision
data = [
    (
        1, "1-InputOutput", "Basic I/O", "Basics", "-", "Java", "Done", "26-Mar-2026",
        "Read integer using Scanner, print with System.out.println()",
        "• import java.util.*;\n• Scanner sc = new Scanner(System.in);\n• sc.nextInt() → int, sc.next() → String, sc.nextLine() → full line",
        "• Forgetting import java.util.*\n• Using sc.nextLine() after sc.nextInt() causes empty string read (buffer issue)\n• println() adds newline, print() does not",
        "No"
    ),
    (
        2, "2-IF_ElseIF", "Conditionals – Grade", "Basics", "-", "Java", "Done", "26-Mar-2026",
        "if / else-if / else ladder to print student grade based on marks",
        "• Conditions checked top-down; first true block runs\n• A≥90, B≥70, C≥50, D≥35, else Fail\n• Always end with else to handle all cases",
        "• Wrong order of conditions (e.g. ≥50 before ≥70) causes wrong grade\n• Missing else → no output for failing marks\n• Using = instead of == for comparison",
        "No"
    ),
    (
        3, "3-Switch", "Switch – Day of Week", "Basics", "-", "Java", "Done", "26-Mar-2026",
        "Switch-case to print day name from day number (1=Monday … 7=Sunday)",
        "• Arrow syntax: case 1 -> System.out.print(\"Monday\");\n• default handles invalid input\n• Arrow syntax does NOT need break",
        "• Old syntax needs break; after each case or fall-through happens\n• Forgetting default case\n• Switch works on int, String, char — NOT float/double",
        "No"
    ),
    (
        4, "4-patterns-1", "Pattern – Solid Rectangle", "Patterns", "-", "Java", "Done", "29-Mar-2026",
        "Print n×m rectangle of stars using nested loops",
        "• Outer loop = rows (i < n)\n• Inner loop = cols (j < m)\n• println() after inner loop for new row",
        "• Using i<n vs i<=n changes row count by 1\n• Forgetting System.out.println() after inner loop\n• Confusing n (rows) and m (cols)",
        "No"
    ),
    (
        5, "5-patterns-2", "Pattern – Right Triangle (stars)", "Patterns", "-", "Java", "Done", "29-Mar-2026",
        "Print right-angled triangle: row i has i stars",
        "• for(i=1; i<=n; i++) → for(j=1; j<=i; j++) print *\n• i controls both row number and star count",
        "• Starting i from 0 prints empty first row\n• Inner loop condition j<=i vs j<i off by one",
        "No"
    ),
    (
        6, "6-patterns-3", "Pattern – Number Triangle", "Patterns", "-", "Java", "Done", "29-Mar-2026",
        "Print triangle with row number repeated (1 / 22 / 333) or column number (1 / 12 / 123)",
        "• Print i for row-number pattern\n• Print j for column-number pattern\n• Toggle between i and j to switch pattern type",
        "• Class name '6-patterns-3' is invalid Java identifier — causes compile error\n• Confusing when to print i vs j",
        "No"
    ),
    (
        7, "7-patterns-4", "Pattern – Decreasing Triangle", "Patterns", "-", "Java", "Done", "29-Mar-2026",
        "Print decreasing star/number triangle; also explored printing j and h counter",
        "• Outer: i from 0 to n; Inner: j from m down to i\n• Use separate counter h++ to print 1,2,3… per row",
        "• Loop bounds: j>i vs j>=i changes last row\n• Resetting h=1 inside outer loop (not outside) is correct",
        "No"
    ),
    (
        8, "8-patterns-5", "Pattern – Same as Rectangle", "Patterns", "-", "Java", "Done", "29-Mar-2026",
        "Solid n×m rectangle (same logic as pattern-1, reinforcement)",
        "• Same as pattern-1\n• Outer i<n, inner j<m",
        "• Duplicate of pattern-1 — ensure you understand the difference before moving on",
        "No"
    ),
    (
        9, "9-patterns-6", "Pattern – Diamond / Expanding Triangle", "Patterns", "-", "Java", "Done", "07-Apr-2026",
        "Print expanding triangle up to 2n rows: row i has i stars (i goes 1 to 2n-1)",
        "• for(i=1; i<n*2; i++) → for(j=0; j<i; j++) print *\n• Creates diamond top-half if combined with shrinking",
        "• Loop i<n*2 vs i<=n*2 changes last row\n• For full diamond need second loop going back down",
        "No"
    ),
    (
        10, "10-patterns-7", "Pattern – Inverted V with dashes", "Patterns", "-", "Java", "Done", "07-Apr-2026",
        "Print inverted triangle of stars with dash padding on sides",
        "• 3 inner loops: dashes(i-1) + stars(n*2-1 - 2*(i-1)) + dashes(i-1)\n• Stars per row = n*2-1-(i-1)*2",
        "• Off-by-one in star count formula is very common here\n• Forgetting the trailing dash loop\n• Loop starts i=1 not i=0",
        "No"
    ),
    (
        11, "11-Math-Reverse-1", "Reverse Integer", "Math", "LC #7", "Java", "Done", "07-Apr-2026",
        "Extract digits with %10, rebuild reversed number with r=r*10+digit, handle overflow",
        "• r = r*10 + (x%10);  x = x/10;\n• Overflow check: if(r > Integer.MAX_VALUE/10 || r < Integer.MIN_VALUE/10) return 0;\n• Works for negative numbers automatically (% keeps sign in Java)",
        "• BUG IN YOUR CODE: overflow check is AFTER multiplication — too late, already overflowed\n• Fix: check BEFORE r=r*10+digit\n• Negative input: Java % keeps sign so -123%10 = -3 ✓",
        "Yes"
    ),
    (
        12, "12-Math-Pallindrom-2", "Palindrome Number", "Math", "LC #9", "Java", "Done", "07-Apr-2026",
        "Reverse the integer and compare with original. Negative numbers → false immediately.",
        "• Store p = x before reversing\n• if(x < 0) return false;\n• Reverse: r=r*10+(x%10); x=x/10;\n• return p == r;",
        "• Forgetting to handle negative numbers\n• Not storing original before reversing\n• Trailing zeros: 10 reversed = 01 = 1 ≠ 10 → correctly false",
        "No"
    ),
    (
        13, "13-Math-Armstrong-3", "Armstrong Number", "Math", "-", "Java", "Done", "07-Apr-2026",
        "Sum of cubes of each digit == original number (for 3-digit numbers)",
        "• s += d*d*d for each digit d\n• Compare s == original\n• Examples: 153=1³+5³+3³ ✓, 371 ✓, 407 ✓",
        "• BUG: function named 'reverse' but checks Armstrong — misleading\n• BUG: hardcoded cube only works for 3-digit; general formula: Math.pow(d, numDigits)\n• numDigits = String.valueOf(n).length()",
        "Yes"
    ),
    (
        14, "14-Recurrsion-Decreasing_Increasing_num-1-2", "Recursion – Print Decreasing / Increasing", "Recursion", "-", "Java", "Done", "08-Apr-2026",
        "Print n down to 1 (decreasing) by printing BEFORE recursive call. Swap print position for increasing.",
        "• Decreasing: print(n) → recurse(n-1)\n• Increasing: recurse(n-1) → print(n)\n• Base case: if(n==0) return;\n• Call stack unwinds for increasing order",
        "• Printing after call = increasing (easy to forget)\n• Base case n==0 not n==1 (else 1 is never printed for decreasing)\n• No return value needed — void recursion",
        "No"
    ),
    (
        15, "15-Recurrsion-fibonacci-5", "Recursion – Fibonacci", "Recursion", "LC #509", "Java", "Done", "08-Apr-2026",
        "fib(n) = fib(n-1) + fib(n-2) with base cases fib(0)=0, fib(1)=1",
        "• Base: if(n==0) return 0; if(n==1) return 1;\n• return fib(n-1) + fib(n-2);\n• Time: O(2^n) — exponential without memoization",
        "• Pure recursion is O(2^n) — TLE for large n on LeetCode\n• Always upgrade to DP memo for n>30\n• fib(0)=0 not 1 — common mistake",
        "No"
    ),
    (
        16, "16-Recurrsion-sum-factorial-3-4", "Recursion – Sum & Factorial", "Recursion", "-", "Java", "Done", "08-Apr-2026",
        "sum(n) = n + sum(n-1). Factorial commented out: n * factorial(n-1)",
        "• Sum base: if(n==1) return 1;\n• Factorial base: if(n==0) return 1;  ← 0! = 1\n• sum(n) = n*(n+1)/2 (math shortcut to verify)",
        "• BUG: factorial base case should be n==0 not n==1 (0! = 1 is undefined with n==1 base)\n• Factorial commented out — finish implementing it\n• Stack overflow for very large n (use iterative for safety)",
        "Yes"
    ),
    (
        17, "17-Recurssion-Array_Reverse-6", "Recursion – Array Reverse (Print)", "Recursion", "-", "Java", "Done", "11-Apr-2026",
        "Print array in reverse using recursion: print arr[n] then recurse with n-1",
        "• print(arr[n]) → recurse(arr, n-1)\n• Base: if(n==0) return arr[0];\n• Does NOT modify array — only prints in reverse",
        "• This only PRINTS reverse, does not reverse in-place\n• In-place reverse: swap(arr[start], arr[end]) + recurse(start+1, end-1)\n• Base for in-place: if(start >= end) return;",
        "Yes"
    ),
    (
        18, "18-Hashing-1", "Hashing – Frequency Count", "Hashing", "-", "Java", "Done", "17-Apr-2026",
        "Count frequency of each element in array using HashMap<Integer,Integer>",
        "• map.put(key, map.getOrDefault(key, 0) + 1);  ← cleaner than containsKey\n• Iterate: for(Map.Entry<K,V> e : map.entrySet())\n• e.getKey(), e.getValue()",
        "• Your code uses containsKey — works but verbose; prefer getOrDefault\n• HashMap is unordered — output order may vary\n• Use LinkedHashMap to preserve insertion order",
        "No"
    ),
    (
        19, "19-dp-fibo-1", "DP – Fibonacci (Memoization + Tabulation)", "DP", "LC #509", "Java", "Done", "17-Apr-2026",
        "Two approaches: top-down memoization (recursion + dp array) and bottom-up tabulation (loop)",
        "• Memo: if(dp[n]!=-1) return dp[n]; ... dp[n]=f(n-1)+f(n-2); return dp[n];\n• Tab:  dp[0]=0; dp[1]=1; for(i=2;i<=n;i++) dp[i]=dp[i-1]+dp[i-2]; return dp[n];\n• Arrays.fill(dp, -1) before memo",
        "• BUG in tabulation: loop runs i<n (exclusive) so dp[n] never filled; returns dp[n-1]\n• Fix: loop i<=n and return dp[n]\n• dp array size must be n+1 not n",
        "Yes"
    ),
    (
        20, "20-dp-climbingstare-2", "DP – Climbing Stairs", "DP", "LC #70", "Java", "Done", "17-Apr-2026",
        "Count ways to climb n stairs (1 or 2 steps at a time). Answer = fib(n+1).",
        "• Base: if(n==0) return 1; if(n<0) return 0;\n• return climb(n-1) + climb(n-2);\n• Memo: dp[n] = climb(n-1,dp) + climb(n-2,dp);\n• Same recurrence as Fibonacci",
        "• BUG 1: 'return=' is invalid Java syntax in climbingStairs_rec — compile error\n• BUG 2: climbingStairs_mem(n) called without dp[] argument in main — compile error\n• Fix: climbingStairs_mem(n, dp)\n• dp array size = n+1",
        "Yes"
    ),
    (
        21, "21-trees-creationand+preorderTraversal", "Tree – Creation + Preorder", "Trees", "-", "Java", "Done", "19-Apr-2026",
        "Build binary tree from int[] using recursion. -1 means null. Preorder traversal: Root→L→R",
        "• static int i = -1; (reset before each call!)\n• i++; if(arr[i]==-1) return null;\n• Node n = new Node(arr[i]); n.left=create(arr); n.right=create(arr);\n• Preorder: print → left → right",
        "• Static i is NOT reset between multiple tree creations — causes wrong tree\n• Array must encode full tree including -1 for every null\n• Preorder ≠ level-order; don't confuse them",
        "No"
    ),
    (
        22, "22-tree-dfs-preorder-inorder-postorder", "Tree – DFS (Pre / In / Post order)", "Trees", "LC #144 #94 #145", "Java", "Done", "19-Apr-2026",
        "Three recursive DFS traversals on binary tree",
        "• Preorder:   Root → Left → Right  (used for tree copy/serialization)\n• Inorder:    Left → Root → Right  (BST → sorted output)\n• Postorder:  Left → Right → Root  (used for tree deletion)\n• Base: if(root==null) return;",
        "• Confusing order of the three traversals\n• Inorder of BST gives sorted array — important property\n• Postorder processes children before parent — used in deletion",
        "No"
    ),
    (
        23, "23-tree-sametree", "Tree – Same Tree", "Trees", "LC #100", "Java", "Done", "19-Apr-2026",
        "Check if two trees are identical by serializing both to lists and comparing",
        "• Direct recursive approach (cleaner):\n  if(p==null && q==null) return true;\n  if(p==null || q==null) return false;\n  return p.val==q.val && isSame(p.left,q.left) && isSame(p.right,q.right);",
        "• BUG: your approach uses two lists + l.add(null) — fragile and memory-heavy\n• l.add(null) works for ArrayList<Integer> but is non-standard\n• Direct recursion is O(n) time O(h) space vs O(n) space for list approach\n• Null handling in list comparison can give false positives",
        "Yes"
    ),
    (
        24, "24-tree-bfs", "Tree – BFS Level Order", "Trees", "LC #102", "Java", "Done", "19-Apr-2026",
        "Level-order traversal using Queue. Null sentinel marks end of each level.",
        "• Queue<Node> q = new LinkedList<>();\n• q.add(root); q.add(null);\n• while(!q.isEmpty()){ n=q.remove(); if(n==null){ println(); if(!q.isEmpty()) q.add(null); } else { print; add children; } }",
        "• BUG: your code adds null back unconditionally → infinite loop when queue has only null left\n• Fix: add null sentinel only if(!q.isEmpty())\n• Also need break/return when queue is empty after removing null\n• Children added only if not null (your code does this correctly ✓)",
        "Yes"
    ),
    (
        25, "25-tree-height", "Tree – Height of Binary Tree", "Trees", "LC #104", "Java", "Done", "19-Apr-2026",
        "Height = max(height(left), height(right)) + 1. Base: null node returns 0.",
        "• if(root==null) return 0;\n• int left=height(root.left); int right=height(root.right);\n• return Math.max(left, right) + 1;\n• Height = number of nodes on longest root-to-leaf path",
        "• Height vs Depth: height counts nodes, depth counts edges (return -1 for null if counting edges)\n• Know which definition the problem uses\n• Balanced tree: |height(left)-height(right)| <= 1",
        "No"
    ),
    (
        26, "26-tree-totalNodes and sum", "Tree – Total Nodes & Sum", "Trees", "-", "Java", "Done", "19-Apr-2026",
        "Count total nodes and sum all node values using post-order recursion",
        "• nodeCount: if(root==null) return 0; return 1+count(left)+count(right);\n• sum:       if(root==null) return 0; return root.data+sum(left)+sum(right);\n• Both follow same post-order pattern",
        "• Both functions have identical structure — remember the pattern\n• Sum can overflow int for large trees — use long\n• Don't forget base case returns 0 (not -1 or null)",
        "No"
    ),
]

# ── write rows ────────────────────────────────────────────────────────────────
for row_idx, row in enumerate(data, 2):
    cat = row[3]
    bg, _ = CAT.get(cat, ("FFFFFF", "000000"))
    rev = row[11]
    for col_idx, val in enumerate(row, 1):
        c = ws.cell(row_idx, col_idx, val)
        # highlight revision needed rows slightly
        cell_bg = "FFE4E1" if rev == "Yes" and col_idx == 12 else bg
        cell_style(c, bg=cell_bg, align="center" if col_idx in (1, 5, 6, 7, 12) else "left",
                   bold=(col_idx == 3))
    ws.row_dimensions[row_idx].height = 90

# ── legend row ────────────────────────────────────────────────────────────────
legend_row = len(data) + 3
ws.cell(legend_row, 1, "Legend:").font = Font(bold=True, size=9)
for i, (cat, (bg, fg)) in enumerate(CAT.items(), 2):
    c = ws.cell(legend_row, i, cat)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal="center")
    c.border = border()

path = r"c:\Users\pranj\OneDrive\Desktop\DSA-(A-Z)\DSA_Notes_v2.xlsx"
wb.save(path)
print("Saved:", path)
